"""Match person project rows to master client references for date enrichment.

Reads the read-only master references workbook and fuzzy-matches person
projects by normalised client + title.  Produces a reviewable artifact and
exposes a lookup function for approved dates.
"""

from __future__ import annotations

import csv
import re
import unicodedata
from difflib import SequenceMatcher
from functools import lru_cache
from pathlib import Path

import pandas as pd

from src.data_loader import PERSON_SHEET_ALIASES, load_projects
from src.runtime_config import get_runtime_config

RUNTIME = get_runtime_config()

# ── Paths ─────────────────────────────────────────────────────────────────────

WORKBOOK_PATH = RUNTIME.workbook_path
MASTER_REFERENCES_PATH = RUNTIME.master_references_path
OVERRIDES_PATH = RUNTIME.data_dir / "project_date_overrides.csv"
REVIEW_OUTPUT_PATH = RUNTIME.outputs_root / "project_date_matches.csv"

# ── Master workbook column names ──────────────────────────────────────────────

_COL_ENGAGEMENT_NUM = "Engagement #"
_COL_COMPANY_NAME = "Company Name"
_COL_ENGAGEMENT_TITLE = "Engagement Title"
_COL_START_DATE = "Start Date"
_COL_END_DATE = "End Date"
_COL_BLC_CONTACT = "BLC Contact"

# ── Thresholds ────────────────────────────────────────────────────────────────

_STRONG_THRESHOLD = 0.78
_POSSIBLE_THRESHOLD = 0.55

# ── Abbreviation expansions ───────────────────────────────────────────────────

_ABBREVIATIONS: dict[str, str] = {
    "ieso": "independent electricity system operator",
    "ciro": "canadian investment regulatory organization",
    "cota": "central ohio transit authority",
    "dioc": "death investigation oversight council",
    "msg": "ministry of the solicitor general",
    "opb": "ontario pension board",
    "oda": "ontario dental association",
    "etr": "express toll route",
    "cic": "crown investment corporation",
    "wsib": "workplace safety and insurance board",
    "oma": "ontario medical association",
    "oeb": "ontario energy board",
    "tpa": "toronto parking authority",
    "tpl": "toronto public library",
    "ttc": "toronto transit commission",
    "yrdsb": "york region district school board",
}

_MUNICIPAL_STOPWORDS = {
    "city",
    "town",
    "township",
    "county",
    "region",
    "regional",
    "municipality",
    "district",
    "village",
    "borough",
    "of",
    "the",
}

# Map BLC Contact field values to canonical person names.
# John Naas / JN = former partner — excluded.
_BLC_CONTACT_MAP: dict[str, str] = {
    "ian shelley": "Ian Shelley",
    "is": "Ian Shelley",
    "jc": "John Connolly",
    "anum": "Anum Nasir",
    "graham": "Graham Pressey",
    "aa": "Abdel Al-Sharif",
    "mt": "Michelle Tam",
    "rl": "",  # unknown
    "vk": "",  # unknown
    "yb": "",  # unknown
}


def _resolve_blc_contacts(raw: str) -> set[str]:
    """Return set of canonical person names from a BLC Contact cell.

    Handles compound values like 'John Naas / JC' or 'IS / AA'.
    Ignores John Naas (former partner).
    """
    names: set[str] = set()
    for part in re.split(r"[/,]", raw):
        part = part.strip().lower()
        if not part or part in ("john naas", "jn"):
            continue
        mapped = _BLC_CONTACT_MAP.get(part, "")
        if mapped:
            names.add(mapped)
    return names


# ── Normalisation helpers ─────────────────────────────────────────────────────


def _normalise(text: str) -> str:
    """Lower-case, strip accents, collapse whitespace, replace & with and."""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = text.lower().strip()
    text = text.replace("&", " and ")
    text = re.sub(r"[^\w\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _expand_abbreviations(text: str) -> str:
    """Expand known acronyms found as standalone tokens."""
    tokens = text.split()
    expanded: list[str] = []
    for token in tokens:
        if token in _ABBREVIATIONS:
            expanded.append(_ABBREVIATIONS[token])
        else:
            expanded.append(token)
    return " ".join(expanded)


def _normalise_client(client: str) -> str:
    """Normalise and expand a client name for comparison."""
    norm = _normalise(client)
    # Expand abbreviations that may appear in parentheses
    # e.g. "Central Ohio Transit Authority (COTA)" -> already has full name
    norm = _expand_abbreviations(norm)
    return norm


def _strip_municipal(client: str) -> str:
    """Remove common municipal words for looser comparison."""
    tokens = client.split()
    return " ".join(t for t in tokens if t not in _MUNICIPAL_STOPWORDS)


def _normalise_title(title: str) -> str:
    """Normalise a project/engagement title for comparison."""
    norm = _normalise(title)
    norm = _expand_abbreviations(norm)
    # Collapse common synonyms
    norm = norm.replace("customer service", "customer experience")
    norm = norm.replace("master plan", "strategic plan")
    norm = re.sub(r"\breview(s)?\b", "review", norm)
    return norm


# ── Scoring ───────────────────────────────────────────────────────────────────


def _sequence_ratio(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a, b).ratio()


def _token_overlap(a: str, b: str) -> float:
    """Jaccard-style overlap on token sets."""
    sa = set(a.split())
    sb = set(b.split())
    if not sa or not sb:
        return 0.0
    return len(sa & sb) / len(sa | sb)


def _score_match(
    person_client: str,
    person_title: str,
    ref_company: str,
    ref_title: str,
    *,
    person_name: str = "",
    blc_contacts: set[str] | None = None,
) -> tuple[float, str]:
    """Return (score, method) for a candidate match.

    Score is in [0, 1].  Higher is better.
    """
    pc = _normalise_client(person_client)
    rc = _normalise_client(ref_company)
    pt = _normalise_title(person_title)
    rt = _normalise_title(ref_title)

    # Client similarity — use the *distinctive* part (stripped municipal words)
    # to avoid "Township of X" matching "Township of Y" on shared prefix alone.
    pc_stripped = _strip_municipal(pc)
    rc_stripped = _strip_municipal(rc)

    if pc_stripped and rc_stripped:
        # Primary: compare distinctive part of client name
        client_core = _sequence_ratio(pc_stripped, rc_stripped)
        # Secondary: full name as a small boost (shared prefix/suffix adds context)
        client_full = _sequence_ratio(pc, rc)
        client_sim = client_core * 0.7 + client_full * 0.3
    else:
        client_core = 1.0  # both empty after stripping → same generic type
        client_sim = _sequence_ratio(pc, rc)

    # Title similarity
    title_seq = _sequence_ratio(pt, rt)
    title_tok = _token_overlap(pt, rt)
    title_sim = max(title_seq, title_tok * 0.95)

    # Combined score — weight client match heavily to avoid cross-client errors
    if client_sim < 0.45:
        # Client names too different — near-zero regardless of title
        score = client_sim * 0.3 + title_sim * 0.1
        method = "low_client"
    else:
        score = client_sim * 0.45 + title_sim * 0.55
        method = "client_title_blend"

    # BLC Contact boost: if the person we're matching IS the BLC Contact
    # on this reference, that's a confirming signal (applied before cap).
    if person_name and blc_contacts and person_name in blc_contacts:
        score = min(score + 0.06, 1.0)
        method += "+blc"

    # Safety cap: if the distinctive part of the client name is weak,
    # prevent a generic-title match from inflating into "strong".
    # Applied last so no boost can override it.
    if client_core < 0.75:
        score = min(score, _STRONG_THRESHOLD - 0.001)
        method += "+capped"

    return round(score, 4), method


# ── Date formatting ───────────────────────────────────────────────────────────


def _format_date(value: object) -> str:
    """Normalise an Excel date cell to YYYY-MM-DD or original string."""
    import datetime

    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, (pd.Timestamp, datetime.datetime)):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    if not text or text.lower() in ("tbd", "nan", "nat"):
        return ""
    if text.lower() == "ongoing":
        return "Ongoing"
    return text


# ── Loading ───────────────────────────────────────────────────────────────────


def _load_master_references() -> list[dict]:
    """Load all rows from the master references workbook."""
    import shutil
    import tempfile

    # Copy to temp file to avoid OneDrive / Excel file locks
    tmp = Path(tempfile.gettempdir()) / "blc_ref_copy.xlsx"
    try:
        shutil.copy2(MASTER_REFERENCES_PATH, tmp)
        read_path = tmp
    except (PermissionError, OSError):
        read_path = MASTER_REFERENCES_PATH

    df = pd.read_excel(
        read_path,
        dtype={_COL_ENGAGEMENT_NUM: str},
    )
    if read_path == tmp:
        tmp.unlink(missing_ok=True)

    rows: list[dict] = []
    for _, row in df.iterrows():
        eng = str(row.get(_COL_ENGAGEMENT_NUM, "")).strip()
        if not eng or eng.lower() in ("n.a.", "na", "nan", ""):
            eng = ""
        company = str(row.get(_COL_COMPANY_NAME, "")).strip()
        title = str(row.get(_COL_ENGAGEMENT_TITLE, "")).strip()
        start = _format_date(row.get(_COL_START_DATE, ""))
        end = _format_date(row.get(_COL_END_DATE, ""))
        blc_raw = str(row.get(_COL_BLC_CONTACT, "")).strip()
        blc_people = _resolve_blc_contacts(blc_raw)
        rows.append(
            {
                "engagement_number": eng,
                "company_name": company,
                "engagement_title": title,
                "start_date": start,
                "end_date": end,
                "blc_contacts": blc_people,
            }
        )
    return rows


def _load_all_person_projects() -> list[dict]:
    """Load all project rows across all people from the extraction workbook."""
    import openpyxl

    wb = openpyxl.load_workbook(str(WORKBOOK_PATH), read_only=True, data_only=True)
    alias_reverse = {v: k for k, v in PERSON_SHEET_ALIASES.items()}
    people: list[str] = []
    for name in wb.sheetnames:
        if name.endswith("_projects"):
            workbook_name = name[: -len("_projects")]
            canonical = alias_reverse.get(workbook_name, workbook_name)
            people.append(canonical)
    wb.close()

    all_rows: list[dict] = []
    for person in people:
        df = load_projects(str(WORKBOOK_PATH), person)
        for _, row in df.iterrows():
            key = str(row.get("Project Key", "")).strip()
            if not key:
                continue
            all_rows.append(
                {
                    "person": person,
                    "project_key": key,
                    "client": str(row.get("Client", "")).strip(),
                    "project_title": str(row.get("Project Title", "")).strip(),
                }
            )
    return all_rows


# ── Core matching ─────────────────────────────────────────────────────────────


def compute_all_matches() -> list[dict]:
    """Match every person project row against master references.

    Returns a list of dicts suitable for CSV export, one per project row.
    """
    refs = _load_master_references()
    projects = _load_all_person_projects()

    results: list[dict] = []
    for proj in projects:
        scored_refs = []

        for ref in refs:
            score, method = _score_match(
                proj["client"],
                proj["project_title"],
                ref["company_name"],
                ref["engagement_title"],
                person_name=proj["person"],
                blc_contacts=ref.get("blc_contacts"),
            )
            if score > 0:
                scored_refs.append((score, method, ref))

        scored_refs.sort(key=lambda x: x[0], reverse=True)

        best_score = scored_refs[0][0] if scored_refs else 0.0
        best_method = scored_refs[0][1] if scored_refs else ""
        best_ref = scored_refs[0][2] if scored_refs else None

        if best_score >= _STRONG_THRESHOLD:
            status = "strong"
        elif best_score >= _POSSIBLE_THRESHOLD:
            status = "possible"
        else:
            status = "unresolved"

        row: dict = {
            "person": proj["person"],
            "project_key": proj["project_key"],
            "client": proj["client"],
            "project_title": proj["project_title"],
            "matched_engagement_number": (
                best_ref["engagement_number"]
                if best_ref and status != "unresolved"
                else ""
            ),
            "matched_company_name": (
                best_ref["company_name"] if best_ref and status != "unresolved" else ""
            ),
            "matched_engagement_title": (
                best_ref["engagement_title"]
                if best_ref and status != "unresolved"
                else ""
            ),
            "start_date": (
                best_ref["start_date"] if best_ref and status != "unresolved" else ""
            ),
            "end_date": (
                best_ref["end_date"] if best_ref and status != "unresolved" else ""
            ),
            "match_status": status,
            "confidence_score": round(best_score, 4),
            "match_method": best_method,
        }

        # Add top 3 alternatives
        alts = [x for x in scored_refs[1:4]]
        for i in range(1, 4):
            idx = i - 1
            if idx < len(alts):
                alt_score, _, alt_ref = alts[idx]
                row[f"alt{i}_engagement_number"] = alt_ref["engagement_number"]
                row[f"alt{i}_company_name"] = alt_ref["company_name"]
                row[f"alt{i}_engagement_title"] = alt_ref["engagement_title"]
                row[f"alt{i}_score"] = round(alt_score, 4)
            else:
                row[f"alt{i}_engagement_number"] = ""
                row[f"alt{i}_company_name"] = ""
                row[f"alt{i}_engagement_title"] = ""
                row[f"alt{i}_score"] = ""

        results.append(row)

    return results


# ── CSV export ────────────────────────────────────────────────────────────────

_REVIEW_COLUMNS = [
    "match_status",
    "person",
    "project_key",
    "client",
    "project_title",
    "matched_engagement_number",
    "matched_company_name",
    "matched_engagement_title",
    "start_date",
    "end_date",
    "confidence_score",
    "match_method",
    "alt1_engagement_number",
    "alt1_company_name",
    "alt1_engagement_title",
    "alt1_score",
    "alt2_engagement_number",
    "alt2_company_name",
    "alt2_engagement_title",
    "alt2_score",
    "alt3_engagement_number",
    "alt3_company_name",
    "alt3_engagement_title",
    "alt3_score",
]


def export_review_csv(path: Path | None = None) -> Path:
    """Generate the reviewable match CSV and return its path."""
    path = path or REVIEW_OUTPUT_PATH
    path.parent.mkdir(parents=True, exist_ok=True)

    matches = compute_all_matches()
    # Sort: strong first, then possible, then unresolved; within each by person
    status_order = {"confirmed": 0, "strong": 1, "possible": 2, "unresolved": 3}
    matches.sort(
        key=lambda r: (
            status_order.get(r["match_status"], 9),
            r["person"],
            r["project_key"],
        )
    )

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=_REVIEW_COLUMNS)
        writer.writeheader()
        writer.writerows(matches)

    return path


# ── Reading back the reviewed CSV ─────────────────────────────────────────────


def _load_reviewed_csv() -> list[dict]:
    """Read the review CSV back, respecting any status edits the user made."""
    if not REVIEW_OUTPUT_PATH.exists():
        return []
    rows: list[dict] = []
    with open(REVIEW_OUTPUT_PATH, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    return rows


# ── Public API ────────────────────────────────────────────────────────────────


@lru_cache(maxsize=1)
def _build_approved_lookup() -> dict[tuple[str, str], dict[str, str]]:
    """Build (person, project_key) -> date dict from confirmed rows in the review CSV."""
    reviewed = _load_reviewed_csv()
    lookup: dict[tuple[str, str], dict[str, str]] = {}

    for row in reviewed:
        status = row.get("match_status", "").strip().lower()
        if status != "confirmed":
            continue
        person = row.get("person", "").strip()
        key = row.get("project_key", "").strip()
        start = row.get("start_date", "").strip()
        end = row.get("end_date", "").strip()
        if not person or not key:
            continue
        if start or end:
            lookup[(person, key)] = {
                "start_date": start,
                "end_date": end,
            }

    # Manual overrides file takes precedence if present
    if OVERRIDES_PATH.exists():
        with open(OVERRIDES_PATH, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                person = row.get("person", "").strip()
                key = row.get("project_key", "").strip()
                start = row.get("approved_start_date", "").strip()
                end = row.get("approved_end_date", "").strip()
                if person and key and (start or end):
                    lookup[(person, key)] = {
                        "start_date": start,
                        "end_date": end,
                    }

    return lookup


def get_approved_dates(person: str, project_key: str) -> dict[str, str]:
    """Return approved start/end dates for a person+project, or empty strings.

    Only returns dates for rows marked 'confirmed' in the review CSV.
    """
    lookup = _build_approved_lookup()
    return lookup.get((person, project_key), {"start_date": "", "end_date": ""})


def clear_cache() -> None:
    """Clear the LRU cache (call after editing the review CSV)."""
    _build_approved_lookup.cache_clear()


# ── Write-back to extraction workbook ─────────────────────────────────────────


def apply_confirmed_to_workbook() -> dict[str, int]:
    """Write engagement numbers from confirmed matches into _key tabs.

    Adds 'Engagement Number' and 'Match Status' columns to each person's
    _key sheet in the extraction workbook.  Only writes for rows whose
    match_status is 'confirmed' in the review CSV.

    Returns a summary dict: {person: count_of_rows_updated}.
    """
    import openpyxl

    reviewed = _load_reviewed_csv()

    # Build lookup: (person, project_key) -> confirmed row data
    confirmed: dict[tuple[str, str], dict] = {}
    for row in reviewed:
        status = row.get("match_status", "").strip().lower()
        if status != "confirmed":
            continue
        person = row.get("person", "").strip()
        key = row.get("project_key", "").strip()
        if person and key:
            confirmed[(person, key)] = row

    if not confirmed:
        return {}

    wb = openpyxl.load_workbook(str(WORKBOOK_PATH))
    alias_reverse = {v: k for k, v in PERSON_SHEET_ALIASES.items()}
    summary: dict[str, int] = {}

    for sheet_name in wb.sheetnames:
        if not sheet_name.endswith("_key"):
            continue

        ws = wb[sheet_name]
        workbook_person = sheet_name[: -len("_key")]
        person = alias_reverse.get(workbook_person, workbook_person)

        # Find or create header columns
        headers = [cell.value for cell in ws[1]]
        key_col_idx = None
        eng_col_idx = None
        status_col_idx = None

        for i, h in enumerate(headers):
            if h == "Project Key":
                key_col_idx = i
            elif h == "Engagement Number":
                eng_col_idx = i
            elif h == "Match Status":
                status_col_idx = i

        if key_col_idx is None:
            continue

        # Add new columns if they don't exist
        if eng_col_idx is None:
            eng_col_idx = len(headers)
            ws.cell(row=1, column=eng_col_idx + 1, value="Engagement Number")
            headers.append("Engagement Number")
        if status_col_idx is None:
            status_col_idx = len(headers)
            ws.cell(row=1, column=status_col_idx + 1, value="Match Status")
            headers.append("Match Status")

        updated = 0
        for row_idx in range(2, ws.max_row + 1):
            project_key = ws.cell(row=row_idx, column=key_col_idx + 1).value
            if not project_key:
                continue
            project_key = str(project_key).strip()

            match = confirmed.get((person, project_key))
            if match:
                eng_num = match.get("matched_engagement_number", "")
                ws.cell(row=row_idx, column=eng_col_idx + 1, value=eng_num)
                ws.cell(row=row_idx, column=status_col_idx + 1, value="confirmed")
                updated += 1

        if updated:
            summary[person] = updated

    if summary:
        wb.save(str(WORKBOOK_PATH))
    wb.close()

    return summary
