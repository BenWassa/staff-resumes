from __future__ import annotations

import re
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.runtime_config import RuntimeConfig


_TIMESTAMPED_NAME_PATTERN = re.compile(r"^(?P<base>.+)_(\d{4}-\d{2}-\d{2}_\d{6})\.xlsx$")


@dataclass(frozen=True)
class GlobalWorkbookResult:
    timestamped_path: Path
    latest_path: Path
    copied_workbooks: int
    copied_sheets: int


def resolve_workbook_path(config: RuntimeConfig, *, refresh_if_configured: bool) -> Path:
    if refresh_if_configured and config.refresh_global_workbook_on_run:
        return build_global_workbook(config).latest_path
    return config.workbook_path


def _iter_person_workbook_files(person_workbooks_dir: Path) -> list[Path]:
    files: list[Path] = []
    for pattern in ("*.xlsx", "*.xlsm"):
        for path in sorted(person_workbooks_dir.glob(pattern)):
            if path.name.startswith("~$"):
                continue
            files.append(path)
    return files


def _safe_sheet_name(name: str, existing: set[str]) -> str:
    candidate = (name or "Sheet").strip()[:31] or "Sheet"
    if candidate not in existing:
        return candidate

    suffix = 1
    while True:
        suffix_text = f"_{suffix}"
        max_base = 31 - len(suffix_text)
        candidate = f"{name[:max_base]}{suffix_text}"
        if candidate not in existing:
            return candidate
        suffix += 1


def _copy_workbook_sheets(source_path: Path, target_wb: Workbook, existing_names: set[str]) -> int:
    source_wb = load_workbook(str(source_path), read_only=True, data_only=False)
    copied_count = 0
    try:
        for sheet_name in source_wb.sheetnames:
            source_ws = source_wb[sheet_name]
            target_sheet_name = _safe_sheet_name(sheet_name, existing_names)
            existing_names.add(target_sheet_name)
            target_ws = target_wb.create_sheet(target_sheet_name)
            for row in source_ws.iter_rows(values_only=True):
                target_ws.append(list(row))
            copied_count += 1
    finally:
        source_wb.close()
    return copied_count


def _cleanup_old_timestamped_files(
    generated_root: Path,
    base_name: str,
    keep_count: int,
    keep_paths: set[Path],
) -> None:
    timestamped_paths: list[tuple[datetime, Path]] = []
    for path in generated_root.glob(f"{base_name}_*.xlsx"):
        match = _TIMESTAMPED_NAME_PATTERN.match(path.name)
        if not match:
            continue
        if match.group("base") != base_name:
            continue
        try:
            ts = datetime.strptime(path.stem[len(base_name) + 1 :], "%Y-%m-%d_%H%M%S")
        except ValueError:
            continue
        timestamped_paths.append((ts, path))

    timestamped_paths.sort(key=lambda item: item[0], reverse=True)
    keep_set = {path.resolve() for _, path in timestamped_paths[:keep_count]}
    keep_set.update(path.resolve() for path in keep_paths)
    for _, path in timestamped_paths:
        if path.resolve() in keep_set:
            continue
        path.unlink(missing_ok=True)


def build_global_workbook(config: RuntimeConfig) -> GlobalWorkbookResult:
    person_workbooks_dir = config.person_workbooks_dir
    generated_root = config.generated_root
    base_name = config.generated_workbook_basename
    latest_path = config.generated_latest_workbook_path

    if not person_workbooks_dir.exists():
        raise FileNotFoundError(
            f"Person workbook folder not found: {person_workbooks_dir}\n"
            f"Please configure this path via the Settings page."
        )

    source_files = _iter_person_workbook_files(person_workbooks_dir)
    if not source_files:
        raise FileNotFoundError(
            f"No person workbook files (*.xlsx, *.xlsm) found in {person_workbooks_dir}"
        )

    generated_root.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    timestamped_path = generated_root / f"{base_name}_{timestamp}.xlsx"
    temp_timestamped_path = generated_root / f".{base_name}_{timestamp}.tmp.xlsx"
    temp_latest_path = generated_root / f".{base_name}_latest.tmp.xlsx"

    target_wb = Workbook()
    default_sheet = target_wb.active
    target_wb.remove(default_sheet)

    existing_sheet_names: set[str] = set()
    copied_sheets = 0
    for source_file in source_files:
        copied_sheets += _copy_workbook_sheets(source_file, target_wb, existing_sheet_names)

    target_wb.save(temp_timestamped_path)
    load_workbook(str(temp_timestamped_path), read_only=True).close()
    temp_timestamped_path.replace(timestamped_path)

    shutil.copy2(timestamped_path, temp_latest_path)
    load_workbook(str(temp_latest_path), read_only=True).close()
    temp_latest_path.replace(latest_path)

    _cleanup_old_timestamped_files(
        generated_root=generated_root,
        base_name=base_name,
        keep_count=max(config.generated_keep_count, 1),
        keep_paths={timestamped_path},
    )

    return GlobalWorkbookResult(
        timestamped_path=timestamped_path,
        latest_path=latest_path,
        copied_workbooks=len(source_files),
        copied_sheets=copied_sheets,
    )
