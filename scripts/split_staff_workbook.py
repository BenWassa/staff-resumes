from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl

GLOBAL_SHEETS = ("README", "Data Dictionary", "HOME")


def find_people(sheet_names: list[str]) -> list[str]:
    """Return person names inferred from `{Name}_projects` tabs."""
    return sorted(
        {
            sheet_name[: -len("_projects")]
            for sheet_name in sheet_names
            if sheet_name.endswith("_projects")
        }
    )


def build_person_workbook(
    source_workbook: Path,
    person_name: str,
    output_path: Path,
    include_global_sheets: tuple[str, ...],
) -> None:
    """Create one workbook containing only global tabs + selected person's tabs."""
    wb = openpyxl.load_workbook(
        filename=source_workbook,
        data_only=False,
        keep_vba=False,
    )

    keep_sheets = set(include_global_sheets)
    keep_sheets.add(f"{person_name}_projects")
    keep_sheets.add(f"{person_name}_profile")

    for sheet_name in list(wb.sheetnames):
        is_key_sheet = sheet_name.endswith("_key")
        if is_key_sheet or sheet_name not in keep_sheets:
            del wb[sheet_name]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()


def export_person_workbooks(
    source_workbook: Path,
    output_dir: Path,
    include_global_sheets: tuple[str, ...],
) -> list[Path]:
    wb = openpyxl.load_workbook(
        filename=source_workbook,
        read_only=True,
        data_only=True,
        keep_vba=True,
    )
    people = find_people(list(wb.sheetnames))
    wb.close()

    output_paths: list[Path] = []
    for person_name in people:
        output_path = output_dir / f"{person_name}.xlsx"
        build_person_workbook(
            source_workbook=source_workbook,
            person_name=person_name,
            output_path=output_path,
            include_global_sheets=include_global_sheets,
        )
        output_paths.append(output_path)

    return output_paths


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Split the Blackline staff workbook into one Excel file per person, "
            "excluding all *_key sheets."
        )
    )
    parser.add_argument(
        "--source",
        type=Path,
        default=Path("data") / "Blackline Staff Project Database.xlsm",
        help="Path to source workbook (.xlsm).",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("outputs") / "person_workbooks",
        help="Folder to write per-person .xlsx files.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    source = args.source.resolve()
    output_dir = args.output_dir.resolve()

    if not source.exists():
        raise FileNotFoundError(f"Source workbook not found: {source}")

    output_paths = export_person_workbooks(
        source_workbook=source,
        output_dir=output_dir,
        include_global_sheets=GLOBAL_SHEETS,
    )

    print(f"Created {len(output_paths)} person workbook(s) in {output_dir}")
    for path in output_paths:
        print(path)


if __name__ == "__main__":
    main()
